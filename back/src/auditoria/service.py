from sqlalchemy.orm import Session, joinedload
from sqlalchemy import and_, or_, func, distinct
from src.auditoria import models as auditoria_models
from src.utilizadores.models import Utilizador
from src.pacientes.models import Paciente
from src.clinica.models import Clinica
from src.marcacoes.models import Marcacao
from src.orcamento.models import Orcamento
from src.faturacao.models import Fatura
from datetime import datetime
from typing import Optional, List
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def listar_auditoria(
    db: Session,
    clinica_id: int,
    skip: int = 0,
    limit: int = 10,
    utilizador_id: Optional[int] = None,
    acao: Optional[str] = None,
    objeto: Optional[str] = None,
    data_inicio: Optional[datetime] = None,
    data_fim: Optional[datetime] = None,
    search: Optional[str] = None
):
    # Base query with eager loading
    query = db.query(auditoria_models.Auditoria).options(
        joinedload(auditoria_models.Auditoria.utilizador),
        joinedload(auditoria_models.Auditoria.clinica)
    )

    # Apply filters
    filters = [auditoria_models.Auditoria.clinica_id == clinica_id]

    if utilizador_id:
        filters.append(auditoria_models.Auditoria.utilizador_id == utilizador_id)

    if acao:
        filters.append(auditoria_models.Auditoria.acao == acao)

    if objeto:
        filters.append(auditoria_models.Auditoria.objeto == objeto)

    if data_inicio:
        filters.append(auditoria_models.Auditoria.data >= data_inicio)

    if data_fim:
        filters.append(auditoria_models.Auditoria.data <= data_fim)

    if search:
        search_term = f"%{search}%"
        filters.append(
            or_(
                auditoria_models.Auditoria.detalhes.ilike(search_term),
                auditoria_models.Auditoria.acao.ilike(search_term),
                auditoria_models.Auditoria.objeto.ilike(search_term)
            )
        )

    if filters:
        query = query.filter(and_(*filters))

    # Get total count for pagination
    total_count = query.count()

    # Apply ordering and pagination
    auditorias = query.order_by(auditoria_models.Auditoria.data.desc()).offset(skip).limit(limit).all()

    # Process results
    respostas = []
    for a in auditorias:
        utilizador_nome = a.utilizador.nome if a.utilizador else None
        clinica_nome = a.clinica.nome if a.clinica else None
        objeto_nome = None

        # Get object name based on type
        if a.objeto == "Utilizador" and a.objeto_id:
            u = db.query(Utilizador).filter_by(id=a.objeto_id).first()
            objeto_nome = u.nome if u else None
        elif a.objeto == "Paciente" and a.objeto_id:
            p = db.query(Paciente).filter_by(id=a.objeto_id).first()
            objeto_nome = p.nome if p else None
        elif a.objeto == "Clinica" and a.objeto_id:
            c = db.query(Clinica).filter_by(id=a.objeto_id).first()
            objeto_nome = c.nome if c else None
        elif a.objeto == "Marcacao" and a.objeto_id:
            m = db.query(Marcacao).filter_by(id=a.objeto_id).first()
            objeto_nome = f"Marcação {m.id}" if m else None
        elif a.objeto == "Orcamento" and a.objeto_id:
            o = db.query(Orcamento).filter_by(id=a.objeto_id).first()
            objeto_nome = f"Orçamento {o.id}" if o else None
        elif a.objeto == "Fatura" and a.objeto_id:
            f = db.query(Fatura).filter_by(id=a.objeto_id).first()
            objeto_nome = f"Fatura {f.id}" if f else None

        respostas.append({
            "id": a.id,
            "utilizador_id": a.utilizador_id,
            "utilizador_nome": utilizador_nome,
            "clinica_id": a.clinica_id,
            "clinica_nome": clinica_nome,
            "acao": a.acao,
            "objeto": a.objeto,
            "objeto_id": a.objeto_id,
            "objeto_nome": objeto_nome,
            "detalhes": a.detalhes,
            "data": a.data,
        })

    return {
        "items": respostas,
        "total": total_count,
        "page": (skip // limit) + 1,
        "per_page": limit,
        "total_pages": (total_count + limit - 1) // limit
    }

def get_auditoria_metadata(db: Session, clinica_id: int):
    """Get metadata for filters including distinct actions, objects, and users"""

    # Get distinct actions for this clinic
    acoes = db.query(distinct(auditoria_models.Auditoria.acao)).filter(
        auditoria_models.Auditoria.acao.isnot(None),
        auditoria_models.Auditoria.clinica_id == clinica_id
    ).all()
    acoes_list = [acao[0] for acao in acoes if acao[0]]

    # Get distinct objects for this clinic
    objetos = db.query(distinct(auditoria_models.Auditoria.objeto)).filter(
        auditoria_models.Auditoria.objeto.isnot(None),
        auditoria_models.Auditoria.clinica_id == clinica_id
    ).all()
    objetos_list = [objeto[0] for objeto in objetos if objeto[0]]

    # Get users who have audit records for this clinic
    utilizadores = db.query(
        Utilizador.id,
        Utilizador.nome
    ).join(
        auditoria_models.Auditoria,
        auditoria_models.Auditoria.utilizador_id == Utilizador.id
    ).filter(
        auditoria_models.Auditoria.clinica_id == clinica_id
    ).distinct().all()

    utilizadores_list = [
        {"id": user.id, "nome": user.nome}
        for user in utilizadores
    ]

    return {
        "acoes": sorted(acoes_list),
        "objetos": sorted(objetos_list),
        "utilizadores": sorted(utilizadores_list, key=lambda x: x["nome"])
    }

def export_auditoria_excel(
    db: Session,
    clinica_id: int,
    filters: Optional[dict] = None,
    selected_ids: Optional[List[int]] = None,
    export_all: bool = False
) -> bytes:
    """Export audit records to Excel format"""

    # Get records based on filters
    if selected_ids:
        # Export only selected records
        query = db.query(auditoria_models.Auditoria).options(
            joinedload(auditoria_models.Auditoria.utilizador),
            joinedload(auditoria_models.Auditoria.clinica)
        ).filter(
            auditoria_models.Auditoria.id.in_(selected_ids),
            auditoria_models.Auditoria.clinica_id == clinica_id
        )
        auditorias = query.order_by(auditoria_models.Auditoria.data.desc()).all()
    elif export_all:
        # Export all records with filters
        if filters:
            # Apply filters (reuse logic from listar_auditoria)
            query = db.query(auditoria_models.Auditoria).options(
                joinedload(auditoria_models.Auditoria.utilizador),
                joinedload(auditoria_models.Auditoria.clinica)
            )

            filter_list = [auditoria_models.Auditoria.clinica_id == clinica_id]
            if filters.get('utilizador_id'):
                filter_list.append(auditoria_models.Auditoria.utilizador_id == filters['utilizador_id'])
            if filters.get('acao'):
                filter_list.append(auditoria_models.Auditoria.acao == filters['acao'])
            if filters.get('objeto'):
                filter_list.append(auditoria_models.Auditoria.objeto == filters['objeto'])
            if filters.get('data_inicio'):
                filter_list.append(auditoria_models.Auditoria.data >= datetime.fromisoformat(filters['data_inicio']))
            if filters.get('data_fim'):
                filter_list.append(auditoria_models.Auditoria.data <= datetime.fromisoformat(filters['data_fim']))
            if filters.get('search'):
                search_term = f"%{filters['search']}%"
                filter_list.append(
                    or_(
                        auditoria_models.Auditoria.detalhes.ilike(search_term),
                        auditoria_models.Auditoria.acao.ilike(search_term),
                        auditoria_models.Auditoria.objeto.ilike(search_term)
                    )
                )

            if filter_list:
                query = query.filter(and_(*filter_list))

            auditorias = query.order_by(auditoria_models.Auditoria.data.desc()).all()
        else:
            # Export all records for this clinic
            auditorias = db.query(auditoria_models.Auditoria).options(
                joinedload(auditoria_models.Auditoria.utilizador),
                joinedload(auditoria_models.Auditoria.clinica)
            ).filter(
                auditoria_models.Auditoria.clinica_id == clinica_id
            ).order_by(auditoria_models.Auditoria.data.desc()).all()
    else:
        # This shouldn't happen, but fallback to empty
        auditorias = []

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Auditoria"

    # Headers
    headers = ["ID", "Data", "Utilizador", "Ação", "Objeto", "Objeto Nome", "Detalhes"]
    ws.append(headers)

    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Add data
    for auditoria in auditorias:
        utilizador_nome = auditoria.utilizador.nome if auditoria.utilizador else f"ID: {auditoria.utilizador_id}"

        # Get object name (reuse logic)
        objeto_nome = None
        if auditoria.objeto == "Utilizador" and auditoria.objeto_id:
            u = db.query(Utilizador).filter_by(id=auditoria.objeto_id).first()
            objeto_nome = u.nome if u else None
        elif auditoria.objeto == "Paciente" and auditoria.objeto_id:
            p = db.query(Paciente).filter_by(id=auditoria.objeto_id).first()
            objeto_nome = p.nome if p else None
        elif auditoria.objeto == "Clinica" and auditoria.objeto_id:
            c = db.query(Clinica).filter_by(id=auditoria.objeto_id).first()
            objeto_nome = c.nome if c else None
        elif auditoria.objeto == "Marcacao" and auditoria.objeto_id:
            m = db.query(Marcacao).filter_by(id=auditoria.objeto_id).first()
            objeto_nome = f"Marcação {m.id}" if m else None
        elif auditoria.objeto == "Orcamento" and auditoria.objeto_id:
            o = db.query(Orcamento).filter_by(id=auditoria.objeto_id).first()
            objeto_nome = f"Orçamento {o.id}" if o else None
        elif auditoria.objeto == "Fatura" and auditoria.objeto_id:
            f = db.query(Fatura).filter_by(id=auditoria.objeto_id).first()
            objeto_nome = f"Fatura {f.id}" if f else None

        row_data = [
            auditoria.id,
            auditoria.data.strftime("%d/%m/%Y %H:%M:%S"),
            utilizador_nome,
            auditoria.acao,
            auditoria.objeto,
            objeto_nome or "",
            auditoria.detalhes or ""
        ]
        ws.append(row_data)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()

def get_auditoria_for_pdf(
    db: Session,
    clinica_id: int,
    filters: Optional[dict] = None,
    selected_ids: Optional[List[int]] = None,
    export_all: bool = False
) -> List[dict]:
    """Get audit records formatted for PDF export"""

    # Get records (reuse logic from Excel export)
    if selected_ids:
        query = db.query(auditoria_models.Auditoria).options(
            joinedload(auditoria_models.Auditoria.utilizador),
            joinedload(auditoria_models.Auditoria.clinica)
        ).filter(
            auditoria_models.Auditoria.id.in_(selected_ids),
            auditoria_models.Auditoria.clinica_id == clinica_id
        )
        auditorias = query.order_by(auditoria_models.Auditoria.data.desc()).all()
    elif export_all:
        if filters:
            query = db.query(auditoria_models.Auditoria).options(
                joinedload(auditoria_models.Auditoria.utilizador),
                joinedload(auditoria_models.Auditoria.clinica)
            )

            filter_list = [auditoria_models.Auditoria.clinica_id == clinica_id]
            if filters.get('utilizador_id'):
                filter_list.append(auditoria_models.Auditoria.utilizador_id == filters['utilizador_id'])
            if filters.get('acao'):
                filter_list.append(auditoria_models.Auditoria.acao == filters['acao'])
            if filters.get('objeto'):
                filter_list.append(auditoria_models.Auditoria.objeto == filters['objeto'])
            if filters.get('data_inicio'):
                filter_list.append(auditoria_models.Auditoria.data >= datetime.fromisoformat(filters['data_inicio']))
            if filters.get('data_fim'):
                filter_list.append(auditoria_models.Auditoria.data <= datetime.fromisoformat(filters['data_fim']))
            if filters.get('search'):
                search_term = f"%{filters['search']}%"
                filter_list.append(
                    or_(
                        auditoria_models.Auditoria.detalhes.ilike(search_term),
                        auditoria_models.Auditoria.acao.ilike(search_term),
                        auditoria_models.Auditoria.objeto.ilike(search_term)
                    )
                )

            if filter_list:
                query = query.filter(and_(*filter_list))

            auditorias = query.order_by(auditoria_models.Auditoria.data.desc()).all()
        else:
            auditorias = db.query(auditoria_models.Auditoria).options(
                joinedload(auditoria_models.Auditoria.utilizador),
                joinedload(auditoria_models.Auditoria.clinica)
            ).filter(
                auditoria_models.Auditoria.clinica_id == clinica_id
            ).order_by(auditoria_models.Auditoria.data.desc()).all()
    else:
        auditorias = []

    # Format for PDF
    formatted_records = []
    for auditoria in auditorias:
        utilizador_nome = auditoria.utilizador.nome if auditoria.utilizador else f"ID: {auditoria.utilizador_id}"

        # Get object name (reuse logic)
        objeto_nome = None
        if auditoria.objeto == "Utilizador" and auditoria.objeto_id:
            u = db.query(Utilizador).filter_by(id=auditoria.objeto_id).first()
            objeto_nome = u.nome if u else None
        elif auditoria.objeto == "Paciente" and auditoria.objeto_id:
            p = db.query(Paciente).filter_by(id=auditoria.objeto_id).first()
            objeto_nome = p.nome if p else None
        elif auditoria.objeto == "Clinica" and auditoria.objeto_id:
            c = db.query(Clinica).filter_by(id=auditoria.objeto_id).first()
            objeto_nome = c.nome if c else None
        elif auditoria.objeto == "Marcacao" and auditoria.objeto_id:
            m = db.query(Marcacao).filter_by(id=auditoria.objeto_id).first()
            objeto_nome = f"Marcação {m.id}" if m else None
        elif auditoria.objeto == "Orcamento" and auditoria.objeto_id:
            o = db.query(Orcamento).filter_by(id=auditoria.objeto_id).first()
            objeto_nome = f"Orçamento {o.id}" if o else None
        elif auditoria.objeto == "Fatura" and auditoria.objeto_id:
            f = db.query(Fatura).filter_by(id=auditoria.objeto_id).first()
            objeto_nome = f"Fatura {f.id}" if f else None

        formatted_records.append({
            "id": auditoria.id,
            "data": auditoria.data.strftime("%d/%m/%Y %H:%M:%S"),
            "utilizador_nome": utilizador_nome,
            "acao": auditoria.acao,
            "objeto": auditoria.objeto,
            "objeto_nome": objeto_nome or "",
            "detalhes": auditoria.detalhes or ""
        })

    return formatted_records